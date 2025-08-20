#!/usr/bin/env python
"""
Prepare bulk user CSV from a single members.xlsx file.

Outputs (in same directory by default):
    bulk_users.csv              -> Clean, deduplicated users ready for bulk_load_users
    invalid_phones.csv          -> Rows skipped due to invalid phone numbers
    generated_ids.csv           -> Mapping of generated student IDs for missing ones
    conflicts.csv               -> Same phone, differing student IDs
    duplicates_phones.csv       -> Raw duplicate occurrences by phone (before dedupe)
    duplicates_names.csv        -> Names appearing more than once (potential duplicates)

Usage:
    python backend/scripts/prepare_bulk_users.py --members backend/members.xlsx \
            --out-dir backend/data_exports

Then run:
  python manage.py bulk_load_users data_exports/bulk_users.csv --send-sms

Assumptions & Rules:
  - All data consolidated into one workbook (members.xlsx).
  - Column names may vary; flexible matching is applied.
  - Phone normalisation:
      * Remove spaces and non-digits.
      * Accept +233XXXXXXXXX or 233XXXXXXXXX -> 0XXXXXXXXX.
      * If 9 digits (missing leading 0) -> prepend 0.
      * If digits length >9 and DOES NOT start with 0 or 233/+233 pattern -> invalid.
      * Final valid format: exactly 10 digits starting with 0.
  - Split full_name into first / last if first/last missing.
  - Duplicate detection by phone (dedup keep first with student_id preference) & by name (report only).

Editable parameters near top: ID_PREFIX, GENERATED_ID_WIDTH
"""
from __future__ import annotations
import argparse
import csv
import sys
from pathlib import Path
from typing import Dict, List, Tuple, Optional, DefaultDict
from collections import defaultdict

try:
    from openpyxl import load_workbook  # Lightweight dependency
except ImportError as e:  # pragma: no cover
    print("openpyxl is required: pip install openpyxl", file=sys.stderr)
    raise

# Config
ID_PREFIX = "GEN"         # Prefix for generated student IDs
GENERATED_ID_WIDTH = 5     # GEN00001, etc.
REQUIRED_OUTPUT_FIELDS = ["student_id", "first_name", "last_name", "phone_number", "email", "program", "year_of_study"]

HEADER_ALIASES = {
    "student_id": {"student id", "student_id", "index number", "index", "id", "index_no", "index no"},
    "first_name": {"first name", "firstname", "first", "given name"},
    "last_name": {"last name", "lastname", "surname", "family name", "last"},
    "phone_number": {"phone", "phone number", "phone_number", "contact", "mobile", "tel"},
    "email": {"email", "email address"},
    "program": {"program", "programme", "course"},
    "year_of_study": {"year", "year of study", "year_of_study", "level", "study year", "class"},
    "full_name": {"full name", "fullname", "name"},
}


def normalize_header(name: str) -> str:
    return name.strip().lower()


def map_headers(raw_headers: List[str]) -> Dict[int, str]:
    mapped: Dict[int, str] = {}
    for idx, h in enumerate(raw_headers):
        h_norm = normalize_header(h)
        target = None
        for canonical, variants in HEADER_ALIASES.items():
            if h_norm in variants:
                target = canonical
                break
        if target:
            mapped[idx] = target
    return mapped


def read_workbook_rows(path: Path) -> List[Dict[str, str]]:
    wb = load_workbook(filename=str(path), data_only=True)
    all_rows: List[Dict[str, str]] = []
    for sheet in wb.worksheets:
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            continue
        headers = [str(h).strip() if h is not None else "" for h in headers]
        header_map = map_headers(headers)
        for row in rows_iter:
            if all((cell is None or str(cell).strip() == "") for cell in row):
                continue  # skip empty
            record: Dict[str, str] = {}
            for col_idx, canonical in header_map.items():
                if col_idx < len(row):
                    val = row[col_idx]
                    if val is None:
                        continue
                    record[canonical] = str(val).strip()
            if record:
                all_rows.append(record)
    return all_rows


def normalize_phone(raw: str) -> Tuple[Optional[str], str]:
    """Return (normalized, reason_if_invalid) applying updated rules.
    Steps:
      1. Strip spaces & non-digits.
      2. Handle +233 / 233 prefix -> local.
      3. If length 9 (no leading 0) -> add 0.
      4. If length > 9 and does not conform to +233/233 patterns and not starting with 0 -> invalid_long_no_prefix.
      5. Result must be 10 digits starting with 0.
    """
    if not raw:
        return None, "empty"
    # Remove spaces & non-digits
    digits = ''.join(ch for ch in raw if ch.isdigit())
    # Country code patterns
    if digits.startswith('233'):
        # Expect 12 total digits for GH (233 + 9 local) else attempt salvage
        if len(digits) >= 12:
            digits = '0' + digits[3:3+9]
        else:
            return None, f"bad_country_length:{digits}"
    # 9 digits missing leading 0
    elif len(digits) == 9:
        digits = '0' + digits
    # If >9 digits but not starting with 0 now -> invalid by rule
    elif len(digits) > 9 and not digits.startswith('0'):
        return None, f"invalid_long_no_prefix:{digits}"
    # Validate final
    if len(digits) != 10 or not digits.startswith('0'):
        return None, f"invalid_format:{digits}"
    return digits, ""


def generate_student_id(counter: int) -> str:
    return f"{ID_PREFIX}{counter:0{GENERATED_ID_WIDTH}d}"


def is_generated_student_id(student_id: str) -> bool:
    return bool(student_id) and student_id.startswith(ID_PREFIX)


def choose_record(existing: Dict[str, str], new: Dict[str, str]) -> Dict[str, str]:
    """Resolve duplicates for same phone.
    Rules:
      1. Prefer real (non-generated) student_id over generated.
      2. Prefer presence of any student_id over none.
      3. If both real and different -> keep existing (conflict tracked separately).
      4. Stable otherwise (keep existing).
    """
    existing_id = existing.get('student_id', '')
    new_id = new.get('student_id', '')
    existing_gen = is_generated_student_id(existing_id)
    new_gen = is_generated_student_id(new_id)

    # If existing has no id and new has any id
    if not existing_id and new_id:
        return new
    # If existing has id and new doesn't
    if existing_id and not new_id:
        return existing
    # Prefer real over generated
    if existing_gen and new_id and not new_gen:
        return new
    if new_gen and existing_id and not existing_gen:
        return existing
    # If both real but different -> keep existing
    return existing


def build_email(username: str, student_id: str) -> str:
    """Build an email; ensure uniqueness by appending student_id fragment if not already present."""
    base = username.lower()
    sid_fragment = ''.join(ch for ch in student_id if ch.isdigit())[-4:]
    if sid_fragment and not base.endswith(sid_fragment):
        base = f"{base}{sid_fragment}"[:30]
    return f"{base}@aamustedbesa.org"


def derive_username(first: str, last: str, student_id: str) -> str:
    base = (first or 'user').lower().replace(' ', '')
    last_part = (last or '').lower().replace(' ', '')[:2]
    if not last_part:
        last_part = ''.join(ch for ch in student_id if ch.isdigit())[:2]
    suffix = ''.join(ch for ch in student_id if ch.isdigit())[-2:]
    return f"{base}_{last_part}{suffix}"[:25]


def sanitize_student_id(raw: str) -> str:
    if not raw:
        return ''
    r = raw.strip()
    # Remove trailing .0 from Excel floats
    if r.endswith('.0') and r.replace('.', '', 1).isdigit():
        r = r[:-2]
    # If it's a float-like string
    if r.count('.') == 1:
        main, dec = r.split('.')
        if dec.isdigit() and int(dec) == 0 and main.isdigit():
            r = main
    return r


def main():
    ap = argparse.ArgumentParser(description="Prepare bulk users CSV from a single Excel source (members.xlsx)")
    ap.add_argument('--members', required=True, help='Path to members.xlsx')
    ap.add_argument('--out-dir', default='.', help='Output directory')
    args = ap.parse_args()

    members_path = Path(args.members)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    combined = read_workbook_rows(members_path)

    users_by_phone: Dict[str, Dict[str, str]] = {}
    invalid_phone_rows: List[Dict[str, str]] = []
    conflicts: List[Dict[str, str]] = []
    generated_ids: List[Tuple[str, str]] = []  # (generated_id, name)
    phone_occurrences: DefaultDict[str, int] = defaultdict(int)
    name_occurrences: DefaultDict[str, List[str]] = defaultdict(list)  # name -> list of phones

    gen_counter = 1

    for row in combined:
        first = row.get('first_name', '').title().strip()
        last = row.get('last_name', '').title().strip()

        # If full_name provided and first/last missing, split it
        if (not first or not last) and row.get('full_name'):
            full = row.get('full_name', '').strip()
            parts = [p for p in full.replace('  ', ' ').split(' ') if p]
            if parts:
                if not first and not last:
                    if len(parts) == 1:
                        first = parts[0].title()
                        last = 'User'
                    else:
                        first = parts[0].title()
                        last = ' '.join(parts[1:]).title()
                elif not first:
                    # last exists, derive first from first token, keep existing last
                    first = parts[0].title()
                elif not last and len(parts) > 1:
                    last = ' '.join(parts[1:]).title()
        student_id = sanitize_student_id(row.get('student_id', ''))
        phone_raw = row.get('phone_number', '').strip()

        phone_norm, err = normalize_phone(phone_raw)
        if not phone_norm:
            row['reason'] = err
            invalid_phone_rows.append(row)
            continue

        phone_occurrences[phone_norm] += 1

        if not student_id:
            student_id = generate_student_id(gen_counter)
            gen_counter += 1
            generated_ids.append((student_id, f"{first} {last}".strip()))

        # Basic username/email if absent
        username = row.get('username') or derive_username(first, last, student_id)
        email = row.get('email') or build_email(username, student_id)

        cleaned = {
            'student_id': student_id,
            'first_name': first or 'Unknown',
            'last_name': last or 'User',
            'phone_number': phone_norm,
            'email': email,
            'program': row.get('program', ''),
            'year_of_study': row.get('year_of_study', ''),
        }

        # Track name occurrences (use cleaned full name)
        full_name_key = f"{first} {last}".strip().lower()
        if full_name_key:
            if phone_norm not in name_occurrences[full_name_key]:
                name_occurrences[full_name_key].append(phone_norm)

        if phone_norm in users_by_phone:
            existing = users_by_phone[phone_norm]
            if (existing['student_id'] != cleaned['student_id'] and
                existing['student_id'] and cleaned['student_id'] and
                not is_generated_student_id(existing['student_id']) and
                not is_generated_student_id(cleaned['student_id'])):
                # Real differing IDs -> log conflict
                conflicts.append({'phone_number': phone_norm, 'existing_student_id': existing['student_id'], 'new_student_id': cleaned['student_id']})
            users_by_phone[phone_norm] = choose_record(existing, cleaned)
        else:
            users_by_phone[phone_norm] = cleaned

    # Write bulk users CSV
    bulk_csv = out_dir / 'bulk_users.csv'
    with bulk_csv.open('w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=REQUIRED_OUTPUT_FIELDS)
        writer.writeheader()
        for rec in users_by_phone.values():
            writer.writerow(rec)

    # Invalid phones
    invalid_csv = out_dir / 'invalid_phones.csv'
    if invalid_phone_rows:
        invalid_headers = sorted({k for r in invalid_phone_rows for k in r.keys()})
        with invalid_csv.open('w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=invalid_headers)
            writer.writeheader()
            writer.writerows(invalid_phone_rows)

    # Generated IDs mapping
    gen_csv = out_dir / 'generated_ids.csv'
    if generated_ids:
        with gen_csv.open('w', newline='', encoding='utf-8') as f:
            w = csv.writer(f)
            w.writerow(['student_id', 'name'])
            w.writerows(generated_ids)

    # Conflicts
    conflicts_csv = out_dir / 'conflicts.csv'
    if conflicts:
        with conflicts_csv.open('w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['phone_number', 'existing_student_id', 'new_student_id'])
            writer.writeheader()
            writer.writerows(conflicts)

    # Duplicate phones (raw occurrences >1)
    dup_phone_csv = out_dir / 'duplicates_phones.csv'
    dup_phone_rows = [ {'phone_number': p, 'count': c} for p, c in phone_occurrences.items() if c > 1 ]
    if dup_phone_rows:
        with dup_phone_csv.open('w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['phone_number', 'count'])
            writer.writeheader()
            writer.writerows(dup_phone_rows)

    # Duplicate names (names appearing more than once -> list phones)
    dup_name_csv = out_dir / 'duplicates_names.csv'
    dup_name_rows = [ {'full_name': n, 'count': len(phones), 'phones': ';'.join(phones)} for n, phones in name_occurrences.items() if len(phones) > 1 ]
    if dup_name_rows:
        with dup_name_csv.open('w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['full_name', 'count', 'phones'])
            writer.writeheader()
            writer.writerows(dup_name_rows)

    print(f"✔ bulk_users.csv written: {bulk_csv}")
    print(f"  Total valid unique users: {len(users_by_phone)}")
    if invalid_phone_rows:
        print(f"⚠ invalid_phones.csv written: {len(invalid_phone_rows)} rows to fix")
    if generated_ids:
        print(f"ℹ generated_ids.csv written: {len(generated_ids)} generated student IDs")
    if conflicts:
        print(f"⚠ conflicts.csv written: {len(conflicts)} phone conflicts require review")
    if dup_phone_rows:
        print(f"ℹ duplicates_phones.csv written: {len(dup_phone_rows)} duplicate phone groups")
    if dup_name_rows:
        print(f"ℹ duplicates_names.csv written: {len(dup_name_rows)} duplicate name groups")


if __name__ == '__main__':
    main()
